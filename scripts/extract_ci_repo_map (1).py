import os
import re
import json
import time
import requests
import pandas as pd
from github import Github
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

# === CONFIGURATION ===
ORG_NAME = "manulife-gwam"
TOKEN = os.getenv("GITHUB_TOKEN")
OUTPUT_JSON_PATH = "data/ci_repo_map_raw.json"
OUTPUT_EXCEL_PATH = "data/ci_repo_map_raw.xlsx"
EXCEL_CI_MAP_PATH = "data/GWAM_github_repo_CI_mapping.xlsx"
CONSOLIDATED_JSON_PATH = "data/consolidated_ci_repo_map.json"
CONSOLIDATED_EXCEL_PATH = "data/consolidated_ci_repo_map.xlsx"
CI_REGEX = re.compile(r"\b(?:CI|APM)\d{4,}\b", re.IGNORECASE)
COMMIT_BRANCH = "develop"
SEARCH_TERMS = ["CI", "APM"]
EXCLUDED_CIS = {"CI000000001234", "CI000000005678"}

# === INITIALIZE GITHUB CLIENT ===
g = Github(TOKEN)
org = g.get_organization(ORG_NAME)
headers = {
    "Authorization": f"token {TOKEN}",
    "Accept": "application/vnd.github+json"
}

def search_code_globally_for_cis():
    ci_repo_map = {}
    seen_items = set()

    for term in SEARCH_TERMS:
        page = 1
        per_page = 100
        query = f"{term}+filename:repo.metadata+org:{ORG_NAME}"

        while True:
            print(f"🔍 Searching GitHub code search: term='{term}', page={page}")
            url = f"https://api.github.com/search/code?q={query}&per_page={per_page}&page={page}"
            resp = requests.get(url, headers=headers)

            if resp.status_code != 200:
                print(f"❌ GitHub search failed: {resp.status_code} {resp.text}")
                break

            data = resp.json()
            items = data.get("items", [])
            if not items:
                break

            for item in items:
                repo_full = item["repository"]["full_name"]
                repo_name = item["repository"]["name"]
                repo = g.get_repo(repo_full)
                default_branch = repo.default_branch
                path = item["path"]

                if (repo_full, path) in seen_items:
                    continue
                seen_items.add((repo_full, path))

                file_url = f"https://raw.githubusercontent.com/{repo_full}/{default_branch}/{path}"
                try:
                    raw = requests.get(file_url, headers=headers).text
                    matches = CI_REGEX.findall(raw)
                    for ci in matches:
                        normalized = ci.upper()
                        if normalized not in EXCLUDED_CIS:
                            ci_repo_map.setdefault(normalized, set()).add(repo_name)
                except Exception as e:
                    print(f"⚠️ Failed to fetch {file_url}: {e}")

            if "next" not in resp.links:
                break

            page += 1
            time.sleep(1)

    for ci in ci_repo_map:
        ci_repo_map[ci] = sorted(list(ci_repo_map[ci]))

    return ci_repo_map

def save_json_and_excel(data):
    os.makedirs(os.path.dirname(OUTPUT_JSON_PATH), exist_ok=True)
    with open(OUTPUT_JSON_PATH, "w") as f:
        json.dump(data, f, indent=2)

    rows = []
    for ci, repos in data.items():
        for repo in repos:
            rows.append({"CI Number": ci, "Repository": repo})

    df = pd.DataFrame(rows)
    df.to_excel(OUTPUT_EXCEL_PATH, index=False)
    print(f"✅ Excel file saved: {OUTPUT_EXCEL_PATH}")

def load_excel_ci_mapping(excel_path):
    df = pd.read_excel(excel_path)
    repo_to_ci = {}
    for _, row in df.iterrows():
        ci_value = str(row.get("Related CI Nums", "")).strip()
        repo_name = str(row.get("Repo Name", "")).strip().split("/")[-1]
        if repo_name:
            repo_to_ci[repo_name] = ci_value.upper() if ci_value.upper() != "N/A" else None
    return repo_to_ci

def merge_mappings(json_map, excel_repo_to_ci):
    combined = defaultdict(set)
    for ci, repos in json_map.items():
        for repo in repos:
            combined[ci.upper()].add(repo.strip())

    for repo, ci in excel_repo_to_ci.items():
        if ci and ci.upper() not in EXCLUDED_CIS:
            combined[ci].add(repo)
        elif not ci:
            for json_ci, repos in json_map.items():
                if repo in repos and json_ci not in EXCLUDED_CIS:
                    combined[json_ci].add(repo)

    return {k: sorted(v) for k, v in combined.items() if k not in EXCLUDED_CIS}

def save_to_excel(consolidated_map):
    wb = Workbook()
    ws = wb.active
    ws.title = "CI to Repo Map"

    ws.append(["CI Number", "Repositories"])
    bold_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_alignment = Alignment(wrap_text=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.border = border

    for ci, repos in consolidated_map.items():
        repo_list = "\n".join(sorted(repos))
        row = [ci, repo_list]
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.column_letter == 'B':
                cell.alignment = wrap_alignment

    for col in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(80, max_length + 5)

    wb.save(CONSOLIDATED_EXCEL_PATH)
    print(f"✅ Excel saved to {CONSOLIDATED_EXCEL_PATH}")

def commit_file_to_repo(file_path, commit_repo_name):
    commit_repo = g.get_repo(commit_repo_name)
    with open(file_path, "rb") as f:
        content = f.read()

    try:
        existing_file = commit_repo.get_contents(file_path, ref=COMMIT_BRANCH)
        commit_repo.update_file(
            path=file_path,
            message=f"Update {os.path.basename(file_path)}",
            content=content,
            sha=existing_file.sha,
            branch=COMMIT_BRANCH
        )
        print(f"✅ Updated file on branch {COMMIT_BRANCH}: {file_path}")
    except Exception:
        commit_repo.create_file(
            path=file_path,
            message=f"Add {os.path.basename(file_path)}",
            content=content,
            branch=COMMIT_BRANCH
        )
        print(f"✅ Created file on branch {COMMIT_BRANCH}: {file_path}")

if __name__ == "__main__":
    commit_repo = os.getenv("COMMIT_REPO")
    if not commit_repo:
        raise ValueError("❌ COMMIT_REPO environment variable not set")

    print("🚀 Starting CI-to-repo scan using GitHub Code Search API...")
    ci_data = search_code_globally_for_cis()
    save_json_and_excel(ci_data)

    print("📥 Loading Excel CI mappings")
    excel_repo_to_ci = load_excel_ci_mapping(EXCEL_CI_MAP_PATH)
    consolidated_map = merge_mappings(ci_data, excel_repo_to_ci)
    save_to_excel(consolidated_map)

    with open(CONSOLIDATED_JSON_PATH, "w") as f:
        json.dump(consolidated_map, f, indent=2)
    print(f"✅ JSON saved to {CONSOLIDATED_JSON_PATH}")

    for path in [OUTPUT_JSON_PATH, OUTPUT_EXCEL_PATH, CONSOLIDATED_JSON_PATH, CONSOLIDATED_EXCEL_PATH]:
        commit_file_to_repo(path, commit_repo)

    print("✅ CI mapping and consolidation complete.")
