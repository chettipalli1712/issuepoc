import os
import json
import pandas as pd
from github import Github
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from collections import defaultdict
import re
import requests
import pytz
import calendar

# === CONFIGURATION ===
TOKEN = os.getenv("GITHUB_TOKEN")
ORG_NAME = "manulife-gwam"
INPUT_JSON = "data/consolidated_ci_repo_map.json"
REPORT_EXCEL = "data/final_ci_report.xlsx"
CI_PATTERN = re.compile(r"\b(?:CI|APM)\d{4,}\b", re.IGNORECASE)

# === HANDLE DATE RANGE BASED ON EVENT ===
event_name = os.getenv("GITHUB_EVENT_NAME", "workflow_dispatch")
event_date = datetime.now(pytz.timezone("US/Eastern"))

if event_name in ["schedule", "workflow_run"]:
    if event_date.day == 16:
        FROM_DATE = event_date.replace(day=1)
        TO_DATE = event_date.replace(day=15)
    else:
        last_day = calendar.monthrange(event_date.year, event_date.month)[1]
        FROM_DATE = event_date.replace(day=1)
        TO_DATE = event_date.replace(day=last_day)
elif event_name == "workflow_dispatch":
    FROM_DATE_str = os.getenv("FROM_DATE")
    TO_DATE_str = os.getenv("TO_DATE")

    if not FROM_DATE_str or not TO_DATE_str:
        raise ValueError("FROM_DATE and TO_DATE must be set to manual trigger")
        
    FROM_DATE = pd.to_datetime(os.getenv("FROM_DATE"), utc=True).tz_convert("US/Eastern")
    TO_DATE = pd.to_datetime(os.getenv("TO_DATE"), utc=True).tz_convert("US/Eastern")
        
else:
    raise ValueError(f"Unsupported event: {event_name}")

# === INIT ===
g = Github(TOKEN)
org = g.get_organization(ORG_NAME)
headers = {
    "Authorization": f"Bearer {TOKEN}",
    "Accept": "application/vnd.github+json"
}

# === HELPER FUNCTIONS ===
def load_ci_repo_map():
    with open(INPUT_JSON, "r") as f:
        return json.load(f)

def get_tags_for_repo(repo_name):
    query = """
    query ($owner: String!, $name: String!, $cursor: String) {
      repository(owner: $owner, name: $name) {
        refs(refPrefix: "refs/tags/", first: 100, after: $cursor, orderBy: {field: TAG_COMMIT_DATE, direction: DESC}) {
          nodes {
            name
            target {
              ... on Tag {
                tagger {
                  date
                }
              }
            }
          }
          pageInfo {
            hasNextPage
            endCursor
          }
        }
      }
    }
    """
    tags = []
    has_next_page = True
    cursor = None

    while has_next_page:
        variables = {
            "owner": ORG_NAME,
            "name": repo_name,
            "cursor": cursor
        }
        response = requests.post(
            "https://api.github.com/graphql",
            json={"query": query, "variables": variables},
            headers=headers
        )
        if response.status_code != 200:
            print(f"❌ Failed to fetch tags for {repo_name}: {response.status_code} {response.text}")
            break

        result = response.json()
        refs = result.get("data", {}).get("repository", {}).get("refs", {})
        for node in refs.get("nodes", []):
            tag_name = node.get("name")
            tag_date = node.get("target", {}).get("tagger", {}).get("date")
            if tag_name and tag_date and (tag_name.startswith("CHG") or tag_name.startswith("INC")):
                tags.append((tag_name, tag_date))

        page_info = refs.get("pageInfo", {})
        has_next_page = page_info.get("hasNextPage", False)
        cursor = page_info.get("endCursor")

    return tags

def build_report(ci_repo_map):
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))

    for ci, repos in ci_repo_map.items():
        for repo in repos:
            tags = get_tags_for_repo(repo)
            for tag_name, tag_date in tags:
                try:
                    tag_match = re.match(r"^(CHG\d+|INC\d+)", tag_name)
                    if not tag_match:
                        continue
                    change_ticket = tag_match.group(1)
                    commit_date = pd.to_datetime(tag_date, utc=True).tz_convert("US/Eastern")
                    if FROM_DATE.date() <= commit_date.date() <= TO_DATE.date():
                        date_str = commit_date.strftime("%Y-%m-%d")
                        grouped[ci][change_ticket][date_str].add(repo)
                except Exception as e:
                    print(f"⚠️ Failed processing tag {tag_name} in {repo}: {e}")

    rows = []
    for ci, change_tickets in grouped.items():
        for ticket, dates in change_tickets.items():
            all_repos = set()
            all_dates = []
            for date, repos in dates.items():
                all_repos.update(repos)
                all_dates.append(date)
            rows.append({
                "CI Number": ci,
                "Change ticket": ticket,
                "Repository": "\n".join(sorted(all_repos)),
                "Date": ", ".join(sorted(all_dates))
            })

    return pd.DataFrame(rows)

def save_report_to_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "CI Tag Report"

    headers = ["CI Number", "Change ticket", "Repository", "Date"]
    ws.append(headers)

    bold_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_alignment = Alignment(wrap_text=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.border = border

    for _, row in df.iterrows():
        values = [row[h] if pd.notna(row[h]) else "NA" for h in headers]
        ws.append(values)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.column_letter == 'C':
                cell.alignment = wrap_alignment

    for col in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(80, max_length + 5)

    wb.save(REPORT_EXCEL)
    print(f"✅ Final report saved to {REPORT_EXCEL}")

# === MAIN ===
if __name__ == "__main__":
    print("📥 Loading consolidated CI-repo mapping...")
    ci_repo_map = load_ci_repo_map()

    print("🔍 Querying tags for all repos within given date range...")
    df_report = build_report(ci_repo_map)

    print("📄 Writing final Excel report...")
    save_report_to_excel(df_report)
    print("✅ Done!")
